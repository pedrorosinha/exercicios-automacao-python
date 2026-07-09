from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

wb = load_workbook('data/output/pivot_table.xlsx')
sheet = wb['Report']

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

bar_chart = BarChart()

data = Reference(sheet, min_col=min_column+1, max_col=max_column, min_row=min_row, max_row=max_row)
categories = Reference(sheet, min_col=min_column, max_col=min_column, min_row=min_row+1, max_row=max_row)

bar_chart.add_data(data, titles_from_data=True)
bar_chart.set_categories(categories)

sheet.add_chart(bar_chart, "B12")

bar_chart.title = "Sales by Product Line"
bar_chart.style = 5

wb.save('data/output/barchart.xlsx')
